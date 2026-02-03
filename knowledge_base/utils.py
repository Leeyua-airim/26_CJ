import os
from typing import List

from openpyxl import load_workbook
from pypdf import PdfReader


def extract_text_from_excel(file_path: str) -> str:
    """
    엑셀(.xlsx/.xls)에서 텍스트를 최대한 범용적으로 추출합니다.

    Parameters:
        file_path (str):
            엑셀 파일 경로
            예: "/tmp/upload.xlsx"

    Returns:
        str:
            시트/행 단위로 합친 텍스트
    """
    wb = load_workbook(filename=file_path, data_only=True)

    lines: List[str] = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        lines.append(f"[SHEET] {sheet_name}")

        for row in ws.iter_rows(values_only=True):
            cells = []

            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                cells.append(s)

            if cells:
                # 행 단위 텍스트
                line = " | ".join(cells)
                lines.append(line)

    return "\n".join(lines)


def extract_text_from_pdf(file_path: str) -> str:
    """
    PDF에서 텍스트를 추출합니다.

    주의:
        - 스캔 PDF(이미지 기반)는 텍스트가 거의 나오지 않을 수 있습니다.
        - Step 4에서는 OCR은 하지 않습니다(정책/복잡도 고려).

    Parameters:
        file_path (str):
            PDF 파일 경로

    Returns:
        str:
            페이지 단위로 합친 텍스트
    """
    reader = PdfReader(file_path)

    pages = []
    page_no = 0
    while page_no < len(reader.pages):
        page = reader.pages[page_no]
        text = page.extract_text() or ""
        text = text.strip()

        if text:
            pages.append(f"[PAGE] {page_no + 1}")
            pages.append(text)

        page_no = page_no + 1

    return "\n".join(pages)


def build_units_from_text(extracted_text: str) -> List[str]:
    """
    문단/행 단위 '기본 유닛'을 만듭니다.

    전략:
        - 줄 단위로 분리
        - 공백/너무 짧은 라인은 제거

    Returns:
        List[str]:
            기본 유닛 리스트
    """
    raw_lines = extracted_text.splitlines()
    units: List[str] = []

    for line in raw_lines:
        s = line.strip()
        if not s:
            continue

        # 너무 짧은 노이즈 제거(필요시 조정)
        if len(s) < 2:
            continue

        units.append(s)

    return units


def chunk_with_context(units: List[str], window: int = 1, max_chars: int = 1200) -> List[str]:
    """
    '문단/행 단위' 유닛에 '주변 문맥'을 붙여 chunk를 생성합니다.

    예:
        window=1이면 i-1, i, i+1을 묶어서 chunk를 만들되,
        max_chars를 넘으면 i만 사용합니다.

    Parameters:
        units (List[str]):
            기본 유닛 리스트
        window (int):
            주변 문맥 포함 범위(앞/뒤)
        max_chars (int):
            chunk 최대 글자 수(간단한 안전장치)

    Returns:
        List[str]:
            chunk 텍스트 리스트
    """
    chunks: List[str] = []

    idx = 0
    while idx < len(units):
        start = idx - window
        end = idx + window

        if start < 0:
            start = 0
        if end >= len(units):
            end = len(units) - 1

        # 후보 chunk
        candidate = "\n".join(units[start : end + 1])

        if len(candidate) > max_chars:
            candidate = units[idx]

        chunks.append(candidate)
        idx = idx + 1

    return chunks


def safe_get_extension(filename: str) -> str:
    """
    파일명에서 확장자를 소문자로 반환합니다.
    """
    _, ext = os.path.splitext(filename)
    return ext.lower().strip()